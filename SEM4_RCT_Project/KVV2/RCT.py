from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import copy

wb = Workbook()
ws = wb.active
ws.title = "Tirzepatide_RCT"

# Header row - same columns as the liraglutide table
headers = [
    "Article #", "Component Type", "Component name", "Component / Placebo",
    "Human / Animal study", "Study group", "Ethnicity / country", "Age (mean)",
    "Age (sd)", "Gender (%)", "Intervention days", "Dosage", "dosage (ml/other)",
    "unit", "Energy from Food", "Macronutrients from Food", "times/day",
    "Lifestyle information if available", "number of timepoints", "Height (m) mean",
    "Height (m) sd", "Unit", "Pre_Weight (kg) mean", "Pre_Weight (kg) sd",
    "Post_Weight (kg) mean", "Post_Weight (kg) sd", "Unit", "Change_Weight (kg) mean",
    "Change_Weight (kg) sd", "Unit", "Pre_BMI (Kg/m^2) mean", "Pre_BMI (Kg/m^2) sd",
    "Post_BMI (Kg/m^2) mean", "Post_BMI (Kg/m^2) sd", "Unit",
    "Pre_BMR (kJ/d) mean", "Pre_BMR (kJ/d) sd", "Post_BMR (kJ/d) mean", "Post_BMR (kJ/d) sd", "Unit",
    "Pre_WC (cm) mean", "Pre_WC (cm) sd", "Post_WC (cm) mean", "Post_WC (cm) sd", "Unit",
    "Pre_HC (cm) mean", "Pre_HC (cm) sd", "Post_HC (cm) mean", "Post_HC (cm) sd", "Unit",
    "Pre_FFM (kg) mean", "Pre_FFM (kg) sd", "Post_FFM (kg) mean", "Post_FFM (kg) sd", "Unit",
    "Pre_Total body fat mass (kg) mean", "Pre_Total body fat mass (kg) sd",
    "Post_Total body fat mass (kg) mean", "Post_Total body fat mass (kg) sd", "Unit",
    "Pre_Total body fat (%) mean", "Pre_Total body fat (%) sd",
    "Post_Total body fat (%) mean", "Post_Total body fat (%) sd", "Unit",
    "Pre_Visceral Fat (%) mean", "Pre_Visceral Fat (%) sd",
    "Post_Visceral Fat (%) mean", "Post_Visceral Fat (%) sd", "Unit",
    "Outcome category", "Outcome name", "Pre_value mean", "Pre_value sd",
    "Post_value mean", "Post_value sd", "Change_value mean", "Change_value sd",
    "Unit_outcome", "ETD", "ETD_95CI_Lower", "ETD_95CI_Upper", "ETD_unit",
    "Favours", "P_value",
    "Pre_SBP (mmHg) mean", "Pre_SBP (mmHg) sd", "Post_SBP (mmHg) mean", "Post_SBP (mmHg) sd", "Unit",
    "Pre_DBP (mmHg) mean", "Pre_DBP (mmHg) sd", "Post_DBP (mmHg) mean", "Post_DBP (mmHg) sd", "Unit",
    "Pre_Glucose (mg/dL) mean", "Pre_Glucose (mg/dL) sd",
    "Post_Glucose (mg/dL) mean", "Post_Glucose (mg/dL) sd", "Unit",
    "Pre_HbA1c (%) mean", "Pre_HbA1c (%) sd", "Post_HbA1c (%) mean", "Post_HbA1c (%) sd", "Unit",
    "Notes", "Active link", "Reference DOI", "No of participants (N)",
    "Type of paper", "Missing parameters"
]

# Style header
header_fill = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
header_font = Font(bold=True, color="FFFFFF", name="Arial", size=9)
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

for col_idx, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align

ws.row_dimensions[1].height = 50

# ---- DATA ----
# Common fields for tirzepatide papers
# Paper 1: Martin et al. 2025 (Nature Medicine) - s41591-025-03774-9
# Paper 2: Heise et al. 2023 (Diabetes Care) - dc221710

# Row template: fill all 103 columns in header order
# We'll define each row as a dict keyed by column header

def make_row(d):
    return [d.get(h, "") for h in headers]

rows = []

# ============================================================
# PAPER 1: Martin 2025 (Nature Medicine) - NCT04311411
# Tirzepatide arm (n=37), 6 weeks
# ============================================================
common_martin_tirz = {
    "Article #": "MA2025",
    "Component Type": "Pharmaceutical",
    "Component name": "Tirzepatide SC",
    "Component / Placebo": "Tirzepatide 5→10 mg",
    "Human / Animal study": "Human",
    "Study group": "Overweight/Obese",
    "Ethnicity / country": "USA",
    "Age (mean)": 44.8,
    "Age (sd)": 10.2,
    "Gender (%)": "F (97%)",
    "Intervention days": 42,
    "Dosage": "5 mg wk 1-3, 10 mg wk 4-6 QW",
    "unit": "mg/day",
    "Lifestyle information if available": "SC QW; dose-escalated 5→10 mg; no dietary/exercise intervention",
    "number of timepoints": 3,
    "Pre_Weight (kg) mean": 97.6,
    "Pre_Weight (kg) sd": 16.0,
    "Pre_BMI (Kg/m^2) mean": 36.1,
    "Pre_BMI (Kg/m^2) sd": 5.7,
    "Pre_WC (cm) mean": 109.8,
    "Pre_WC (cm) sd": 12.2,
    "Reference DOI": "10.1038/s41591-025-03774-9",
    "Active link": "https://doi.org/10.1038/s41591-025-03774-9",
    "No of participants (N)": 37,
    "Type of paper": "Parallel RCT",
}

# Row 1: Energy intake at lunch week 3 (primary)
r = dict(common_martin_tirz)
r.update({
    "Outcome category": "Energy intake",
    "Outcome name": "Ad libitum lunch energy intake – week 3 (primary)",
    "Pre_value mean": 926.4, "Pre_value sd": 371.2,
    "Change_value mean": -532.4, "Unit_outcome": "kcal",
    "ETD": -524.6, "ETD_95CI_Lower": -648.1, "ETD_95CI_Upper": -401.0, "ETD_unit": "kcal",
    "Favours": "Tirzepatide", "P_value": "<0.0001",
    "Notes": "Primary outcome; ANCOVA ETD vs placebo; -59.1% median reduction",
    "Change_Weight (kg) mean": -3.7,
})
rows.append(make_row(r))

# Row 2: Energy intake at lunch week 6
r = dict(common_martin_tirz)
r.update({
    "Outcome category": "Energy intake",
    "Outcome name": "Ad libitum lunch energy intake – week 6",
    "Pre_value mean": 926.4, "Pre_value sd": 371.2,
    "Change_value mean": -657.8, "Unit_outcome": "kcal",
    "ETD": -686.0, "ETD_95CI_Lower": -830.7, "ETD_95CI_Upper": -541.4, "ETD_unit": "kcal",
    "Favours": "Tirzepatide", "P_value": "<0.0001",
    "Notes": "-72.4% median reduction at week 6; sustained reduction vs placebo and liraglutide",
    "Change_Weight (kg) mean": -7.0,
})
rows.append(make_row(r))

# Row 3: Energy intake vs liraglutide week 3
r = dict(common_martin_tirz)
r.update({
    "Outcome category": "Energy intake",
    "Outcome name": "Ad libitum lunch energy intake – week 3 vs liraglutide",
    "Change_value mean": -532.4, "Unit_outcome": "kcal",
    "ETD": -233.1, "ETD_95CI_Lower": -358.0, "ETD_95CI_Upper": -108.3, "ETD_unit": "kcal",
    "Favours": "Tirzepatide", "P_value": "0.0004",
    "Notes": "Tirzepatide superior to liraglutide at week 3 for energy intake",
})
rows.append(make_row(r))

# Row 4: Body weight week 3
r = dict(common_martin_tirz)
r.update({
    "Outcome category": "Body weight",
    "Outcome name": "Body weight change from baseline – week 3",
    "Pre_value mean": 98.2, "Pre_value sd": 16.2,
    "Change_value mean": -3.7, "Unit_outcome": "kg",
    "ETD": -3.3, "ETD_95CI_Lower": -3.8, "ETD_95CI_Upper": -2.7, "ETD_unit": "kg",
    "Favours": "Tirzepatide", "P_value": "<0.0001",
    "Notes": "Significant vs placebo (-0.5 kg) and liraglutide (-1.9 kg) at week 3",
})
rows.append(make_row(r))

# Row 5: Body weight week 6
r = dict(common_martin_tirz)
r.update({
    "Outcome category": "Body weight",
    "Outcome name": "Body weight change from baseline – week 6",
    "Pre_value mean": 98.2, "Pre_value sd": 16.2,
    "Post_Weight (kg) mean": 91.2,
    "Change_value mean": -7.0, "Unit_outcome": "kg",
    "ETD": -6.5, "ETD_95CI_Lower": -7.4, "ETD_95CI_Upper": -5.6, "ETD_unit": "kg",
    "Favours": "Tirzepatide", "P_value": "<0.0001",
    "Notes": "Greatest weight reduction of all groups; -3.4 kg vs liraglutide (p<0.0001)",
})
rows.append(make_row(r))

# Row 6: Fasting overall appetite VAS week 3
r = dict(common_martin_tirz)
r.update({
    "Outcome category": "Appetite (VAS)",
    "Outcome name": "Fasting overall appetite VAS score – week 3",
    "Pre_value mean": 25.3, "Pre_value sd": 14.0,
    "Change_value mean": 22.7, "Unit_outcome": "VAS 0-100 (higher=less appetite)",
    "ETD": 20.6, "ETD_95CI_Lower": 12.1, "ETD_95CI_Upper": 29.0, "ETD_unit": "VAS score",
    "Favours": "Tirzepatide", "P_value": "<0.0001",
    "Notes": "Significant vs placebo (+2.2) and liraglutide (+8.8) at week 3",
})
rows.append(make_row(r))

# Row 7: Fasting overall appetite VAS week 6
r = dict(common_martin_tirz)
r.update({
    "Outcome category": "Appetite (VAS)",
    "Outcome name": "Fasting overall appetite VAS score – week 6",
    "Pre_value mean": 25.3, "Pre_value sd": 14.0,
    "Change_value mean": 28.4, "Unit_outcome": "VAS 0-100 (higher=less appetite)",
    "ETD": 30.8, "ETD_95CI_Lower": 21.3, "ETD_95CI_Upper": 40.4, "ETD_unit": "VAS score",
    "Favours": "Tirzepatide", "P_value": "<0.0001",
    "Notes": "Continued improvement at week 6; ETD vs liraglutide 18.2 (p=0.0002)",
})
rows.append(make_row(r))

# Row 8: Food Craving Inventory (FCI) overall week 3
r = dict(common_martin_tirz)
r.update({
    "Outcome category": "Control of eating",
    "Outcome name": "Food Craving Inventory (FCI) overall score – week 3",
    "Pre_value mean": 2.4, "Pre_value sd": 0.6,
    "Change_value mean": -0.7, "Unit_outcome": "score",
    "ETD": -0.4, "ETD_95CI_Lower": -0.7, "ETD_95CI_Upper": -0.2, "ETD_unit": "score",
    "Favours": "Tirzepatide", "P_value": "0.0008",
    "Notes": "Reduced cravings for high-fats, sweets, carbs, fast-food fats (not fruits/veg)",
})
rows.append(make_row(r))

# Row 9: FCQ-S overall week 3
r = dict(common_martin_tirz)
r.update({
    "Outcome category": "Control of eating",
    "Outcome name": "Food Craving Questionnaire-State (FCQ-S) overall score – week 3",
    "Pre_value mean": 2.9, "Pre_value sd": 0.8,
    "Change_value mean": -1.0, "Unit_outcome": "score",
    "ETD": -0.9, "ETD_95CI_Lower": -1.2, "ETD_95CI_Upper": -0.6, "ETD_unit": "score",
    "Favours": "Tirzepatide", "P_value": "<0.0001",
    "Notes": "All subscales decreased: hunger, lack of control, anticipation, desire",
})
rows.append(make_row(r))

# Row 10: Power of Food Scale (PFS) week 3
r = dict(common_martin_tirz)
r.update({
    "Outcome category": "Control of eating",
    "Outcome name": "Power of Food Scale (PFS) overall score – week 3",
    "Pre_value mean": 3.2, "Pre_value sd": 0.9,
    "Change_value mean": -1.0, "Unit_outcome": "score",
    "ETD": -1.0, "ETD_95CI_Lower": -1.3, "ETD_95CI_Upper": -0.7, "ETD_unit": "score",
    "Favours": "Tirzepatide", "P_value": "<0.0001",
    "Notes": "All subscales (food available, present, tasted) reduced; less responsiveness to food environment",
})
rows.append(make_row(r))

# Row 11: Eating Inventory disinhibition week 3
r = dict(common_martin_tirz)
r.update({
    "Outcome category": "Control of eating",
    "Outcome name": "Eating Inventory disinhibition (tendency to overeat) – week 3",
    "Pre_value mean": 9.5, "Pre_value sd": 3.5,
    "Change_value mean": -3.5, "Unit_outcome": "score",
    "ETD": -3.1, "ETD_95CI_Lower": -4.5, "ETD_95CI_Upper": -1.8, "ETD_unit": "score",
    "Favours": "Tirzepatide", "P_value": "<0.0001",
    "Notes": "Greater reduction in tendency to overeat vs placebo; also significant vs liraglutide (ETD -2.3, p=0.0014)",
})
rows.append(make_row(r))

# Row 12: Eating Inventory perceived hunger week 3
r = dict(common_martin_tirz)
r.update({
    "Outcome category": "Appetite (VAS)",
    "Outcome name": "Eating Inventory perceived hunger – week 3",
    "Pre_value mean": 7.1, "Pre_value sd": 4.3,
    "Change_value mean": -3.6, "Unit_outcome": "score",
    "ETD": -3.2, "ETD_95CI_Lower": -4.6, "ETD_95CI_Upper": -1.8, "ETD_unit": "score",
    "Favours": "Tirzepatide", "P_value": "<0.0001",
    "Notes": "Significant vs placebo and liraglutide (ETD -1.9, p=0.0071) at week 3",
})
rows.append(make_row(r))

# Row 13: Cognitive restraint week 3 (non-significant)
r = dict(common_martin_tirz)
r.update({
    "Outcome category": "Control of eating",
    "Outcome name": "Eating Inventory cognitive restraint – week 3",
    "Pre_value mean": 9.3, "Pre_value sd": 4.6,
    "Change_value mean": 1.6, "Unit_outcome": "score",
    "ETD": 1.3, "ETD_95CI_Lower": -0.1, "ETD_95CI_Upper": 2.8, "ETD_unit": "score",
    "P_value": "0.0682",
    "Notes": "Non-significant vs placebo; tirzepatide did not affect volitional cognitive restraint",
})
rows.append(make_row(r))

# Row 14: Barratt Impulsiveness Scale week 3
r = dict(common_martin_tirz)
r.update({
    "Outcome category": "Control of eating",
    "Outcome name": "Barratt Impulsiveness Scale (BIS) total score – week 3",
    "Pre_value mean": 2.3, "Pre_value sd": 0.2,
    "Change_value mean": -0.02, "Unit_outcome": "score",
    "ETD": -0.10, "ETD_95CI_Lower": -0.17, "ETD_95CI_Upper": -0.02, "ETD_unit": "score",
    "Favours": "Tirzepatide", "P_value": "0.0098",
    "Notes": "Reduced impulsiveness vs placebo; cognitive instability and attentional impulsiveness decreased",
})
rows.append(make_row(r))

# Row 15: BOLD fMRI high-fat/high-sugar - medial frontal gyrus week 3
r = dict(common_martin_tirz)
r.update({
    "Outcome category": "Neuroimaging (fMRI)",
    "Outcome name": "BOLD fMRI – high-fat/high-sugar food activation – medial frontal gyrus – week 3",
    "Unit_outcome": "BOLD parameter",
    "ETD": -0.082, "ETD_unit": "BOLD parameter",
    "Favours": "Tirzepatide", "P_value": "0.0335",
    "Notes": "Decreased activation to FoodHiF/HiS in medial frontal gyrus, cingulate gyrus (p=0.0306), hippocampus (p=0.0221), orbitofrontal cortex (p=0.0321) vs placebo at week 3",
})
rows.append(make_row(r))

# ============================================================
# Placebo arm - Martin 2025
# ============================================================
common_martin_plac = {
    "Article #": "MA2025",
    "Component Type": "Pharmaceutical",
    "Component name": "Placebo",
    "Component / Placebo": "Placebo SC",
    "Human / Animal study": "Human",
    "Study group": "Overweight/Obese",
    "Ethnicity / country": "USA",
    "Age (mean)": 46.2,
    "Age (sd)": 9.5,
    "Gender (%)": "F (92%)",
    "Intervention days": 42,
    "Dosage": "QW",
    "unit": "",
    "Lifestyle information if available": "No dietary/exercise intervention",
    "number of timepoints": 3,
    "Pre_Weight (kg) mean": 98.7,
    "Pre_Weight (kg) sd": 20.5,
    "Pre_BMI (Kg/m^2) mean": 36.2,
    "Pre_BMI (Kg/m^2) sd": 5.9,
    "Pre_WC (cm) mean": 110.3,
    "Pre_WC (cm) sd": 14.3,
    "Reference DOI": "10.1038/s41591-025-03774-9",
    "Active link": "https://doi.org/10.1038/s41591-025-03774-9",
    "No of participants (N)": 39,
    "Type of paper": "Parallel RCT",
}

r = dict(common_martin_plac)
r.update({
    "Outcome category": "Energy intake",
    "Outcome name": "Ad libitum lunch energy intake – week 3",
    "Pre_value mean": 893.1, "Pre_value sd": 384.2,
    "Change_value mean": -7.9, "Unit_outcome": "kcal",
    "Notes": "Reference arm; minimal change from baseline (-7.9 kcal, +1.9%)",
})
rows.append(make_row(r))

r = dict(common_martin_plac)
r.update({
    "Outcome category": "Body weight",
    "Outcome name": "Body weight change from baseline – week 6",
    "Change_value mean": -0.6, "Unit_outcome": "kg",
    "Notes": "Reference arm; essentially weight stable over 6 weeks",
})
rows.append(make_row(r))

r = dict(common_martin_plac)
r.update({
    "Outcome category": "Appetite (VAS)",
    "Outcome name": "Fasting overall appetite VAS score – week 3",
    "Pre_value mean": 26.3, "Pre_value sd": 11.3,
    "Change_value mean": 2.2, "Unit_outcome": "VAS 0-100",
    "Notes": "Reference arm; no significant change in appetite",
})
rows.append(make_row(r))

# ============================================================
# Liraglutide arm - Martin 2025
# ============================================================
common_martin_lira = {
    "Article #": "MA2025",
    "Component Type": "Pharmaceutical",
    "Component name": "Liraglutide SC",
    "Component / Placebo": "Liraglutide 3.0 mg",
    "Human / Animal study": "Human",
    "Study group": "Overweight/Obese",
    "Ethnicity / country": "USA",
    "Age (mean)": 43.7,
    "Age (sd)": 11.9,
    "Gender (%)": "F (66%)",
    "Intervention days": 42,
    "Dosage": "3 mg/day QD",
    "unit": "mg/day",
    "Lifestyle information if available": "Dose-escalated 0.6→1.2→1.8→2.4→3.0 mg QD; no dietary/exercise intervention",
    "number of timepoints": 3,
    "Pre_Weight (kg) mean": 101.1,
    "Pre_Weight (kg) sd": 19.5,
    "Pre_BMI (Kg/m^2) mean": 36.2,
    "Pre_BMI (Kg/m^2) sd": 5.4,
    "Pre_WC (cm) mean": 110.7,
    "Pre_WC (cm) sd": 14.5,
    "Reference DOI": "10.1038/s41591-025-03774-9",
    "Active link": "https://doi.org/10.1038/s41591-025-03774-9",
    "No of participants (N)": 38,
    "Type of paper": "Parallel RCT",
}

r = dict(common_martin_lira)
r.update({
    "Outcome category": "Energy intake",
    "Outcome name": "Ad libitum lunch energy intake – week 3",
    "Pre_value mean": 1044.8, "Pre_value sd": 562.3,
    "Change_value mean": -299.3, "Unit_outcome": "kcal",
    "ETD": -291.4, "ETD_95CI_Lower": -412.1, "ETD_95CI_Upper": -170.8,
    "Favours": "Liraglutide", "P_value": "<0.0001",
    "Notes": "Active comparator arm; -31.5% median reduction at week 3; inferior to tirzepatide",
})
rows.append(make_row(r))

r = dict(common_martin_lira)
r.update({
    "Outcome category": "Energy intake",
    "Outcome name": "Ad libitum lunch energy intake – week 6",
    "Pre_value mean": 1044.8, "Pre_value sd": 562.3,
    "Change_value mean": -314.5, "Unit_outcome": "kcal",
    "ETD": -342.8, "ETD_95CI_Lower": -483.6, "ETD_95CI_Upper": -202.0,
    "Favours": "Liraglutide", "P_value": "<0.0001",
    "Notes": "-28.8% median reduction at week 6",
})
rows.append(make_row(r))

r = dict(common_martin_lira)
r.update({
    "Outcome category": "Body weight",
    "Outcome name": "Body weight change from baseline – week 6",
    "Change_value mean": -3.7, "Unit_outcome": "kg",
    "ETD": -3.1, "ETD_95CI_Lower": -4.0, "ETD_95CI_Upper": -2.3,
    "Favours": "Liraglutide", "P_value": "<0.0001",
    "Notes": "Significant vs placebo but less than tirzepatide (-7.0 kg)",
})
rows.append(make_row(r))

r = dict(common_martin_lira)
r.update({
    "Outcome category": "Appetite (VAS)",
    "Outcome name": "Fasting overall appetite VAS score – week 6",
    "Pre_value mean": 25.6, "Pre_value sd": 18.7,
    "Change_value mean": 10.2, "Unit_outcome": "VAS 0-100",
    "ETD": 12.6, "ETD_95CI_Lower": 3.4, "ETD_95CI_Upper": 21.8,
    "Favours": "Liraglutide", "P_value": "0.0080",
    "Notes": "Significant vs placebo at week 6 only (not week 3); less effect than tirzepatide",
})
rows.append(make_row(r))

# ============================================================
# PAPER 2: Heise 2023 (Diabetes Care) - dc22-1710
# Tirzepatide 15 mg arm, 28 weeks, T2D
# ============================================================
common_heise_tirz = {
    "Article #": "HE2023",
    "Component Type": "Pharmaceutical",
    "Component name": "Tirzepatide SC",
    "Component / Placebo": "Tirzepatide 15 mg",
    "Human / Animal study": "Human",
    "Study group": "T2D/Obese",
    "Ethnicity / country": "Germany (Profil)",
    "Age (mean)": 61.1,
    "Age (sd)": 7.1,
    "Gender (%)": "M (69%)",
    "Intervention days": 196,
    "Dosage": "15 mg/week QW",
    "unit": "mg/week",
    "Lifestyle information if available": "Dose-escalated 2.5→15 mg SC QW; study-wide dietary background",
    "number of timepoints": 4,
    "Pre_Weight (kg) mean": 94.2,
    "Pre_Weight (kg) sd": 14.0,
    "Pre_BMI (Kg/m^2) mean": 31.3,
    "Pre_BMI (Kg/m^2) sd": 5.0,
    "Pre_WC (cm) mean": 113.5,
    "Pre_WC (cm) sd": 8.9,
    "Pre_Total body fat mass (kg) mean": 36.8,
    "Pre_Total body fat mass (kg) sd": 11.5,
    "Pre_FFM (kg) mean": 57.7,
    "Pre_FFM (kg) sd": 9.3,
    "Pre_HbA1c (%) mean": 7.8,
    "Pre_HbA1c (%) sd": 0.7,
    "Pre_Glucose (mg/dL) mean": 139.3,
    "Pre_Glucose (mg/dL) sd": 30.2,
    "Reference DOI": "10.2337/dc22-1710",
    "Active link": "https://doi.org/10.2337/dc22-1710",
    "No of participants (N)": 45,
    "Type of paper": "Parallel RCT",
}

# Body weight week 28
r = dict(common_heise_tirz)
r.update({
    "Outcome category": "Body weight",
    "Outcome name": "Body weight change from baseline – week 28",
    "Pre_value mean": 94.2, "Pre_value sd": 14.0,
    "Post_value mean": 83.4,
    "Change_value mean": -11.2, "Unit_outcome": "kg",
    "ETD": -11.2, "ETD_95CI_Lower": -14.0, "ETD_95CI_Upper": -8.4, "ETD_unit": "kg",
    "Favours": "Tirzepatide", "P_value": "<0.001",
    "Notes": "Significant vs placebo and vs semaglutide 1 mg (-6.9 kg; ETD -4.3 kg, p<0.001)",
    "Change_Weight (kg) mean": -11.2,
    "Post_Weight (kg) mean": 83.4,
})
rows.append(make_row(r))

# Fat mass week 28
r = dict(common_heise_tirz)
r.update({
    "Outcome category": "Body composition",
    "Outcome name": "Total fat mass change from baseline – week 28",
    "Pre_Total body fat mass (kg) mean": 36.8, "Pre_Total body fat mass (kg) sd": 11.5,
    "Post_Total body fat mass (kg) mean": 27.1,
    "Pre_value mean": 36.8, "Pre_value sd": 11.5,
    "Post_value mean": 27.1,
    "Change_value mean": -9.7, "Unit_outcome": "kg",
    "ETD": -9.6, "ETD_95CI_Lower": -12.4, "ETD_95CI_Upper": -6.9, "ETD_unit": "kg",
    "Favours": "Tirzepatide", "P_value": "<0.001",
    "Notes": "-26.5% fat mass reduction; significantly greater than semaglutide (-5.9 kg, -16.1%); ETD vs semaglutide -3.8 kg (p=0.002)",
})
rows.append(make_row(r))

# Fat-free mass week 28
r = dict(common_heise_tirz)
r.update({
    "Outcome category": "Body composition",
    "Outcome name": "Fat-free mass change from baseline – week 28",
    "Pre_FFM (kg) mean": 57.7, "Pre_FFM (kg) sd": 9.3,
    "Post_FFM (kg) mean": 56.1,
    "Pre_value mean": 57.7, "Pre_value sd": 9.3,
    "Post_value mean": 56.1,
    "Change_value mean": -1.6, "Unit_outcome": "kg",
    "ETD": -1.5, "ETD_95CI_Lower": -2.3, "ETD_95CI_Upper": -0.7, "ETD_unit": "kg",
    "Favours": "Tirzepatide", "P_value": "<0.001",
    "Notes": "-2.8% lean mass reduction; also significant vs semaglutide (-0.8 kg, p=0.018); weight loss predominantly fat mass",
})
rows.append(make_row(r))

# Fasting appetite week 28
r = dict(common_heise_tirz)
r.update({
    "Outcome category": "Appetite (VAS)",
    "Outcome name": "Fasting overall appetite VAS score – week 28",
    "Unit_outcome": "VAS 0-100 (higher=less appetite)",
    "ETD": 15.0, "ETD_95CI_Lower": 4.1, "ETD_95CI_Upper": 25.9, "ETD_unit": "VAS score",
    "Favours": "Tirzepatide", "P_value": "0.007",
    "Notes": "Significant vs placebo; not significant vs semaglutide (ETD 5.3, p=0.260); both active treatments reduced appetite similarly",
})
rows.append(make_row(r))

# Energy intake week 28
r = dict(common_heise_tirz)
r.update({
    "Outcome category": "Energy intake",
    "Outcome name": "Ad libitum buffet-style lunch energy intake – week 28",
    "Pre_value mean": 1105.0, "Pre_value sd": 343.7,
    "Change_value mean": -348.4, "Unit_outcome": "kcal",
    "ETD": -309.8, "ETD_95CI_Lower": -423.0, "ETD_95CI_Upper": -196.6, "ETD_unit": "kcal",
    "Favours": "Tirzepatide", "P_value": "<0.001",
    "Notes": "Significant vs placebo; numerically greater than semaglutide (64 kcal difference) but not statistically significant (p=0.187)",
})
rows.append(make_row(r))

# ============================================================
# Semaglutide arm - Heise 2023
# ============================================================
common_heise_sema = {
    "Article #": "HE2023",
    "Component Type": "Pharmaceutical",
    "Component name": "Semaglutide SC",
    "Component / Placebo": "Semaglutide 1 mg",
    "Human / Animal study": "Human",
    "Study group": "T2D/Obese",
    "Ethnicity / country": "Germany (Profil)",
    "Age (mean)": 63.7,
    "Age (sd)": 5.9,
    "Gender (%)": "M (77%)",
    "Intervention days": 196,
    "Dosage": "1 mg/week QW",
    "unit": "mg/week",
    "Lifestyle information if available": "Dose-escalated 0.25→1 mg SC QW; active GLP-1 RA comparator",
    "number of timepoints": 4,
    "Pre_Weight (kg) mean": 92.7,
    "Pre_Weight (kg) sd": 14.0,
    "Pre_BMI (Kg/m^2) mean": 30.8,
    "Pre_BMI (Kg/m^2) sd": 3.8,
    "Pre_WC (cm) mean": 109.7,
    "Pre_WC (cm) sd": 9.2,
    "Pre_Total body fat mass (kg) mean": 35.3,
    "Pre_Total body fat mass (kg) sd": 8.0,
    "Pre_FFM (kg) mean": 56.3,
    "Pre_FFM (kg) sd": 10.3,
    "Pre_HbA1c (%) mean": 7.7,
    "Pre_HbA1c (%) sd": 0.6,
    "Pre_Glucose (mg/dL) mean": 128.6,
    "Pre_Glucose (mg/dL) sd": 25.0,
    "Reference DOI": "10.2337/dc22-1710",
    "Active link": "https://doi.org/10.2337/dc22-1710",
    "No of participants (N)": 44,
    "Type of paper": "Parallel RCT",
}

r = dict(common_heise_sema)
r.update({
    "Outcome category": "Body weight",
    "Outcome name": "Body weight change from baseline – week 28",
    "Pre_value mean": 92.7, "Pre_value sd": 14.0,
    "Post_value mean": 87.7,
    "Change_value mean": -6.9, "Unit_outcome": "kg",
    "Notes": "Active comparator; less weight loss than tirzepatide (-11.2 kg); ETD vs tirzepatide +4.3 kg",
    "Change_Weight (kg) mean": -6.9,
})
rows.append(make_row(r))

r = dict(common_heise_sema)
r.update({
    "Outcome category": "Body composition",
    "Outcome name": "Total fat mass change from baseline – week 28",
    "Pre_value mean": 35.3, "Pre_value sd": 8.0,
    "Change_value mean": -5.9, "Unit_outcome": "kg",
    "Notes": "Fat mass -16.1%; significantly less than tirzepatide (-26.5%)",
})
rows.append(make_row(r))

r = dict(common_heise_sema)
r.update({
    "Outcome category": "Energy intake",
    "Outcome name": "Ad libitum buffet-style lunch energy intake – week 28",
    "Pre_value mean": 1131.0, "Pre_value sd": 375.6,
    "Change_value mean": -284.1, "Unit_outcome": "kcal",
    "ETD": -245.5, "ETD_95CI_Lower": -377.0, "ETD_95CI_Upper": -114.0,
    "Favours": "Semaglutide", "P_value": "<0.001",
    "Notes": "Significant vs placebo; similar to tirzepatide (not statistically different, p=0.187)",
})
rows.append(make_row(r))

# Placebo arm - Heise 2023
common_heise_plac = {
    "Article #": "HE2023",
    "Component Type": "Pharmaceutical",
    "Component name": "Placebo",
    "Component / Placebo": "Placebo SC",
    "Human / Animal study": "Human",
    "Study group": "T2D/Obese",
    "Ethnicity / country": "Germany (Profil)",
    "Age (mean)": 60.4,
    "Age (sd)": 7.6,
    "Gender (%)": "M (75%)",
    "Intervention days": 196,
    "Dosage": "QW",
    "unit": "",
    "number of timepoints": 4,
    "Pre_Weight (kg) mean": 98.7,
    "Pre_Weight (kg) sd": 14.6,
    "Pre_BMI (Kg/m^2) mean": 32.2,
    "Pre_BMI (Kg/m^2) sd": 4.0,
    "Pre_WC (cm) mean": 109.2,
    "Pre_WC (cm) sd": 12.0,
    "Pre_HbA1c (%) mean": 7.9,
    "Pre_HbA1c (%) sd": 0.5,
    "Pre_Glucose (mg/dL) mean": 126.6,
    "Pre_Glucose (mg/dL) sd": 23.6,
    "Reference DOI": "10.2337/dc22-1710",
    "Active link": "https://doi.org/10.2337/dc22-1710",
    "No of participants (N)": 28,
    "Type of paper": "Parallel RCT",
}

r = dict(common_heise_plac)
r.update({
    "Outcome category": "Body weight",
    "Outcome name": "Body weight change from baseline – week 28",
    "Change_value mean": 0, "Unit_outcome": "kg",
    "Notes": "Reference arm; weight-stable over 28 weeks",
})
rows.append(make_row(r))

r = dict(common_heise_plac)
r.update({
    "Outcome category": "Energy intake",
    "Outcome name": "Ad libitum buffet-style lunch energy intake – week 28",
    "Pre_value mean": 1252.7, "Pre_value sd": 483.2,
    "Change_value mean": -38.6, "Unit_outcome": "kcal",
    "Notes": "Reference arm; minimal reduction in energy intake",
})
rows.append(make_row(r))

# Write rows
fill_tirz = PatternFill("solid", start_color="DDEEFF", end_color="DDEEFF")
fill_plac = PatternFill("solid", start_color="F5F5F5", end_color="F5F5F5")
fill_lira = PatternFill("solid", start_color="FFF3CD", end_color="FFF3CD")
fill_sema = PatternFill("solid", start_color="E8F5E9", end_color="E8F5E9")

article_fills = {
    "MA2025_Tirzepatide SC": fill_tirz,
    "MA2025_Placebo": fill_plac,
    "MA2025_Liraglutide SC": fill_lira,
    "HE2023_Tirzepatide SC": fill_tirz,
    "HE2023_Semaglutide SC": fill_sema,
    "HE2023_Placebo": fill_plac,
}

thin = Side(style='thin', color='CCCCCC')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for row_idx, row_data in enumerate(rows, 2):
    key = f"{row_data[0]}_{row_data[2]}"
    fill = article_fills.get(key, PatternFill())
    for col_idx, val in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font = Font(name="Arial", size=9)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = border
        if fill:
            cell.fill = fill

# Column widths
ws.column_dimensions['A'].width = 10
ws.column_dimensions['B'].width = 14
ws.column_dimensions['C'].width = 18
ws.column_dimensions['D'].width = 22
for col in range(5, len(headers)+1):
    ws.column_dimensions[get_column_letter(col)].width = 14

# Freeze top row
ws.freeze_panes = "A2"

# Add legend sheet
ws2 = wb.create_sheet("Legend")
legend_data = [
    ["Color", "Meaning"],
    ["Blue", "Tirzepatide arm"],
    ["Yellow/Cream", "Liraglutide arm (active comparator in MA2025)"],
    ["Green", "Semaglutide arm (active comparator in HE2023)"],
    ["Grey", "Placebo arm"],
]
for r_idx, row in enumerate(legend_data, 1):
    for c_idx, val in enumerate(row, 1):
        cell = ws2.cell(row=r_idx, column=c_idx, value=val)
        cell.font = Font(bold=(r_idx==1), name="Arial")

fills_leg = [None, fill_tirz, fill_lira, fill_sema, fill_plac]
for i, f in enumerate(fills_leg):
    if f and i > 0:
        ws2.cell(row=i+1, column=1).fill = f

ws2.cell(row=1,column=1).value = "Article #: MA2025 = Martin et al. 2025 (Nature Medicine, doi:10.1038/s41591-025-03774-9)"
ws2.cell(row=2,column=1).value = "Article #: HE2023 = Heise et al. 2023 (Diabetes Care, doi:10.2337/dc22-1710)"
ws2.cell(row=3,column=1).value = "Both studies funded by Eli Lilly and Company (manufacturer of tirzepatide/Mounjaro/Zepbound)"
ws2.cell(row=4,column=1).value = "Color coding: Blue=Tirzepatide | Yellow=Liraglutide comparator | Green=Semaglutide comparator | Grey=Placebo"
ws2.column_dimensions['A'].width = 90

wb.save('/home/ibab/tirzepatide_RCT_table.xlsx')
print("Done")
